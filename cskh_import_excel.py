import argparse
import json
import os
import re
import secrets
import string
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from dotenv import load_dotenv
from rapidfuzz import fuzz, process
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine
from unidecode import unidecode


def _env(name: str, default: Optional[str] = None) -> str:
    v = os.getenv(name, default)
    if v is None or v == "":
        raise RuntimeError(f"Missing env var: {name}")
    return v


def _normalize_text(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    if not s:
        return ""
    s = unidecode(s)
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def normalize_vn_phone(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    digit_only = re.sub(r"\D+", "", s)
    if digit_only.startswith("84") and len(digit_only) >= 10:
        digit_only = "0" + digit_only[2:]
    # Một ô có thể chứa nhiều SĐT (vd "0356...; 0964...") — chuẩn hóa match dùng số đầu tiên.
    found = re.findall(r"0\d{9}", digit_only)
    if found:
        return found[0]
    return digit_only


def _collapse_digits_to_vn10(raw_digits: str) -> Optional[str]:
    """Chuỗi chỉ gồm chữ số -> một SĐT trong nước 10 số (0xxxxxxxxx) nếu hợp lệ."""
    if not raw_digits:
        return None
    d = raw_digits
    if d.startswith("84") and len(d) >= 11:
        d = "0" + d[2:13]
    if len(d) == 9 and d[0] in "35789":
        d = "0" + d
    if len(d) >= 10 and d[0] == "0":
        cand = d[:10]
        if re.fullmatch(r"0\d{9}", cand):
            return cand
    return None


def _read_spaced_local_phone(s: str, start: int) -> Optional[Tuple[int, str]]:
    """
    Từ vị trí `start` (ký tự '0'), đọc đủ 10 chữ số (cho phép khoảng trắng/gạch/giữa các số).
    Trả về (index_ký_tự_sau_cùng_của_SĐT, phone10) hoặc None.
    """
    if start >= len(s) or s[start] != "0":
        return None
    j = start
    digits: List[str] = []
    while j < len(s) and len(digits) < 10:
        if s[j].isdigit():
            digits.append(s[j])
            j += 1
        elif s[j] in " \t\u00a0.-+/" and digits:
            j += 1
        else:
            break
    if len(digits) != 10:
        return None
    phone = "".join(digits)
    if not re.fullmatch(r"0\d{9}", phone):
        return None
    return j, phone


def _read_plus84_phone(s: str, start: int) -> Optional[Tuple[int, str]]:
    """Đọc +84... / 084... -> SĐT 10 số trong nước."""
    if s.startswith("+84", start):
        j = start + 3
    elif s.startswith("084", start):
        j = start + 3
    else:
        return None
    while j < len(s) and s[j] in " \t\u00a0-":
        j += 1
    d: List[str] = []
    while j < len(s) and len(d) < 11:
        if s[j].isdigit():
            d.append(s[j])
        elif s[j] in " \t\u00a0.-/" and d:
            pass
        else:
            break
        j += 1
    raw = "".join(d)
    phone = _collapse_digits_to_vn10(raw)
    if not phone:
        return None
    return j, phone


def _read_nine_digit_mobile(s: str, start: int) -> Optional[Tuple[int, str]]:
    """9 chữ số liền, bắt đầu 3/5/7/8/9 — thường là thiếu số 0 đầu (vd 948283559)."""
    if start + 9 > len(s):
        return None
    if start > 0 and s[start - 1] == "0":
        return None
    if not re.match(r"^[35789]\d{8}$", s[start : start + 9]):
        return None
    if start > 0 and s[start - 1].isdigit():
        return None
    if start + 9 < len(s) and s[start + 9].isdigit():
        return None
    return start + 9, "0" + s[start : start + 9]


def find_all_vn_phone_spans(s: str) -> List[Tuple[int, int, str]]:
    """Các đoạn (start, end, phone10) trong chuỗi, trái sang phải, không chồng lấn."""
    out: List[Tuple[int, int, str]] = []
    i = 0
    n = len(s)
    while i < n:
        hit: Optional[Tuple[int, int, str]] = None
        if i + 3 <= n and (s.startswith("+84", i) or s.startswith("084", i)):
            r84 = _read_plus84_phone(s, i)
            if r84:
                end_j, ph = r84
                hit = (i, end_j, ph)
        elif s[i] == "0" and i + 1 < n and s[i + 1].isdigit():
            r0 = _read_spaced_local_phone(s, i)
            if r0:
                end_j, ph = r0
                hit = (i, end_j, ph)
        elif i + 9 <= n:
            r9 = _read_nine_digit_mobile(s, i)
            if r9:
                end_j, ph = r9
                hit = (i, end_j, ph)
        if hit:
            st, en, ph = hit
            out.append((st, en, ph))
            i = en
            continue
        i += 1
    return out


def preprocess_customer_name_cell_for_parse(s: str) -> str:
    """Chuẩn bị chuỗi: gộp khoảng trắng; bỏ ngoặc bao quanh SĐT cuối ô (vd \"Tên ( 0365... )\")."""
    t = (s or "").strip()
    if not t:
        return ""
    t = re.sub(r"\s+", " ", t)
    # "Name ( 0365 889 168 )" hoặc "(+84 911 ...)" ở cuối -> đưa SĐT ra ngoài ngoặc
    t = re.sub(
        r"\(\s*((?:\+84|084|0)[\d\s\.\-+]{8,24})\s*\)\s*$",
        lambda m: " " + m.group(1).replace(" ", " ").strip() + " ",
        t,
    )
    return t.strip()


def split_combined_customer_cell(combined: str) -> Tuple[str, str]:
    """
    Tách (tên, sđt) khi cuối chuỗi là SĐT VN 10 số (bắt đầu 0), phía trước có khoảng trắng/gạch ngang.
    Ví dụ: "Phạm Văn Long 0394789468" -> ("Phạm Văn Long", "0394789468")
    """
    s = preprocess_customer_name_cell_for_parse(combined or "")
    if not s:
        return "", ""
    m = re.match(r"^(.*)[\s\u00a0\-–—]+(0[\d\s\.\-]{8,22})\s*$", s)
    if m:
        tail = re.sub(r"\D", "", m.group(2))
        phone = _collapse_digits_to_vn10(tail)
        if phone:
            name = m.group(1).strip()
            return name, phone
    m2 = re.match(r"^(.*)[\s\u00a0\-–—]+(\+84[\d\s\.\-+]{8,24})\s*$", s)
    if m2:
        tail = re.sub(r"\D", "", m2.group(2))
        if tail.startswith("84"):
            tail = "0" + tail[2:]
        phone = _collapse_digits_to_vn10(tail[:12])
        if phone:
            return m2.group(1).strip(), phone
    return s, ""


def parse_contacts_from_combined_cell(combined: str) -> Tuple[List[Tuple[str, str]], str]:
    """
    Tách nhiều cặp (tên, SĐT) và phần đuôi không gắn SĐt (đưa vào notes khi import).

    Quét SĐT: +84, 0xxxx (có khoảng giữa các số), 9 số thiếu 0 (không nằm sau chữ 0).
    Fallback: tách theo `/` từng đoạn có SĐT ở cuối.
    """
    s0 = preprocess_customer_name_cell_for_parse(combined or "")
    if not s0:
        return [], ""

    spans = find_all_vn_phone_spans(s0)
    if spans:
        contacts: List[Tuple[str, str]] = []
        cursor = 0
        for st, en, ph in spans:
            chunk = s0[cursor:st]
            chunk = re.sub(r"[\s/|,:;–—\-]+$", "", chunk)
            chunk = re.sub(r"^[\s/|,:;–—\-]+", "", chunk).strip()
            contacts.append((chunk, ph))
            cursor = en
        tail = s0[cursor:].strip()
        tail = re.sub(r"^[\s\-–—/+|]+", "", tail)
        fixed: List[Tuple[str, str]] = []
        for nm, p in contacts:
            if not nm.strip() and p:
                fixed.append(("", p))
            else:
                fixed.append((nm.strip(), p))
        if len(fixed) >= 2 and all(not n for n, _ in fixed):
            phones_only = [p for _, p in fixed if p]
            if phones_only:
                return [( "", p) for p in phones_only], tail
        return fixed, tail

    parts = [p.strip() for p in re.split(r"\s*/\s*", s0) if p.strip()]
    if not parts:
        return [], ""
    out: List[Tuple[str, str]] = []
    for p in parts:
        n, ph = split_combined_customer_cell(p)
        if ph:
            out.append((n, ph))
        else:
            out.append((p.strip(), ""))
    return out, ""


def _tail_is_extra_person_name_not_description(tail: str) -> bool:
    """Đuôi sau SĐT cuối: nếu giống tên người thì gộp vào tên KH; nếu giống mô tả dịch vụ thì để notes."""
    t = (tail or "").strip()
    if not t or re.search(r"\d", t):
        return False
    if len(t) > 48:
        return False
    n = _normalize_text(t)
    if any(
        x in n
        for x in (
            "ho tro",
            "may loc",
            "hotro",
            "mayloc",
            "ho tro may",
            "xu ly",
            "bao hanh",
        )
    ):
        return False
    if len(t.split()) > 8:
        return False
    return True


def apply_customer_column_split_for_import(customer_name: str, phone_raw: Any) -> Tuple[str, Any, str]:
    """
    Chuẩn hóa cặp (tên, sđt) kiểu legacy: cột C dính số (nhiều người `/`, `-`, +84, ngoặc, SĐT có khoảng trắng…), cột D có thể trống.
    - Quét mọi SĐT trong ô; gộp tên; đuôi giống tên phụ (không chứa số, không giống mô tả \"hỗ trợ/máy lọc\") gộp vào tên.
    - D trống: ghi các SĐT vào D cách \"; \" (normalize_vn_phone lấy SĐT đầu cho DB).
    - D đã có: giữ D; phần còn lại có thể ghi notes.
    """
    c = str_cell(customer_name)
    d_plain = str_cell(phone_raw)
    contacts, tail_remainder = parse_contacts_from_combined_cell(c)
    if not contacts:
        return c, phone_raw, ""

    names_joined = " / ".join(n for n, _ in contacts if n)
    phones_known = [p for _, p in contacts if p and len(p) == 10]
    if not names_joined and phones_known:
        names_joined = " / ".join(phones_known)

    tr = tail_remainder.strip() if tail_remainder else ""
    if tr and _tail_is_extra_person_name_not_description(tr):
        names_joined = f"{names_joined} / {tr}".strip(" /") if names_joined else tr
        tr = ""

    note_bits: List[str] = []
    if tr:
        note_bits.append(f"Mô tả/ghi chú kèm trong ô: {tr}")

    if d_plain:
        name_out = names_joined if phones_known else c
        extras: List[str] = []
        for n, p in contacts[1:]:
            if p:
                extras.append(f"{n} — {p}" if n else p)
            elif n:
                extras.append(n)
        if extras:
            note_bits.append("Thêm từ cột khách: " + "; ".join(extras))
        extra = "\n".join(note_bits) if note_bits else ""
        return name_out, phone_raw, extra

    if not phones_known:
        extra = "\n".join(note_bits) if note_bits else ""
        return (names_joined or c), phone_raw, extra

    phone_str: str = phones_known[0] if len(phones_known) == 1 else "; ".join(phones_known)
    extra = "\n".join(note_bits) if note_bits else ""
    return (names_joined or c), phone_str, extra


def normalize_excel_customer_columns(
    excel_path: str,
    sheet_name: Optional[str],
    out_path: str,
    all_sheets: bool,
    also_merge_created_at: bool = True,
) -> None:
    """
    Ghi file Excel mới: tách tên/SĐT ở cột khách hàng (legacy C hoặc cskh_customers.name),
    điền cột SĐT (legacy D / cskh_customer_phones.phone) khi đang trống.
    Mặc định sau đó gộp luôn cột ngày tạo trên cùng file out (logic merge như --merge-created-at-out).
    """
    xls = pd.ExcelFile(excel_path)
    target_sheets = list(xls.sheet_names) if all_sheets else [sheet_name or xls.sheet_names[0]]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sn in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sn, header=0).copy()
            if sn not in target_sheets:
                df.to_excel(writer, sheet_name=sn, index=False)
                continue

            col_map, header_mode = build_column_map(df)
            j_name = col_map.get("customer_name") if header_mode else 2
            j_phone = col_map.get("phone") if header_mode else 3

            if header_mode and (j_name is None or j_phone is None):
                print(f"Sheet {sn!r}: không map được customer_name/phone — giữ nguyên.")
                df.to_excel(writer, sheet_name=sn, index=False)
                continue
            if not header_mode and df.shape[1] < 4:
                print(f"Sheet {sn!r}: ít hơn 4 cột — giữ nguyên.")
                df.to_excel(writer, sheet_name=sn, index=False)
                continue

            for i in range(len(df)):
                orig_c = df.iat[i, j_name]
                orig_d = df.iat[i, j_phone] if j_phone < df.shape[1] else None
                c0 = str_cell(orig_c)
                new_c, new_d, _ = apply_customer_column_split_for_import(c0, orig_d)
                df.iat[i, j_name] = new_c
                df.iat[i, j_phone] = new_d

            df.to_excel(writer, sheet_name=sn, index=False)

    if also_merge_created_at:
        merge_excel_created_at_column(
            excel_path=out_path,
            sheet_name=sheet_name,
            out_path=out_path,
            all_sheets=all_sheets,
        )


def _detect_created_at_column_1based(ws: Any) -> int:
    """Cột Excel 1-based: tiêu đề `cskh_tickets.created_at` / `tickets.created_at`; không thấy thì cột A (1)."""
    mc = int(ws.max_column or 1)
    for c in range(1, mc + 1):
        lab = _header_label(ws.cell(row=1, column=c).value)
        if classify_excel_header(lab) == "created_at":
            return c
    return 1


def _merge_date_key_for_cells(v: Any) -> Optional[str]:
    """Khóa so sánh để gộp các dòng cùng \"ngày hiệu lực\" sau ffill."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, pd.Timestamp):
        if pd.isna(v):
            return None
        return pd.Timestamp(v).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            from openpyxl.utils.datetime import from_excel

            dt = from_excel(float(v))
            if isinstance(dt, datetime):
                return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return None
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:19] if len(s) >= 19 else (s[:10] + " 00:00:00")
    return s


def _unmerge_ranges_touching_column(ws: Any, col_idx: int) -> None:
    for mcr in list(ws.merged_cells.ranges):
        if mcr.min_col <= col_idx <= mcr.max_col:
            ws.unmerge_cells(str(mcr))


def merge_excel_created_at_column(
    excel_path: str,
    sheet_name: Optional[str],
    out_path: str,
    all_sheets: bool,
) -> None:
    """
    Gộp ô theo chiều dọc ở cột ngày tạo (created_at): cùng giá trị sau ffill (ô đầu block có ngày,
    các dòng dưới trống = cùng ngày) -> merge thành một ô (giống merge thủ công trên Excel).
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    xls = pd.ExcelFile(excel_path)
    target_sheets = set(xls.sheet_names) if all_sheets else {sheet_name or xls.sheet_names[0]}

    wb = load_workbook(excel_path)
    for sn in wb.sheetnames:
        if sn not in target_sheets:
            continue
        ws = wb[sn]
        col = _detect_created_at_column_1based(ws)
        max_row = int(ws.max_row or 1)
        if max_row < 2:
            print(f"Sheet {sn!r}: không có dòng dữ liệu — bỏ qua gộp.")
            continue

        _unmerge_ranges_touching_column(ws, col)

        vals: List[Any] = [ws.cell(row=r, column=col).value for r in range(2, max_row + 1)]
        last: Any = None
        effective: List[Any] = []
        for v in vals:
            if v is not None and not (isinstance(v, float) and pd.isna(v)):
                s = str(v).strip()
                if s and s.lower() != "nan":
                    last = v
            effective.append(last)

        i = 0
        n = len(effective)
        merged_blocks = 0
        while i < n:
            if effective[i] is None:
                i += 1
                continue
            key = _merge_date_key_for_cells(effective[i])
            if key is None:
                i += 1
                continue
            j = i + 1
            while j < n and effective[j] is not None and _merge_date_key_for_cells(effective[j]) == key:
                j += 1
            excel_start = i + 2
            excel_end = (j - 1) + 2
            if excel_end > excel_start:
                first_val: Any = None
                for r in range(excel_start, excel_end + 1):
                    cv = ws.cell(row=r, column=col).value
                    if cv is not None and not (isinstance(cv, float) and pd.isna(cv)):
                        ts = str(cv).strip()
                        if ts and ts.lower() != "nan":
                            first_val = cv
                            break
                if first_val is None:
                    first_val = effective[i]
                for r in range(excel_start + 1, excel_end + 1):
                    ws.cell(row=r, column=col).value = None
                top = ws.cell(row=excel_start, column=col)
                top.value = first_val
                top.alignment = Alignment(vertical="center", wrap_text=True)
                ws.merge_cells(
                    start_row=excel_start,
                    start_column=col,
                    end_row=excel_end,
                    end_column=col,
                )
                merged_blocks += 1
            i = j
        print(f"Sheet {sn!r}: cột {col} — đã gộp {merged_blocks} khối ngày (logic ffill giống import).")

    wb.save(out_path)


def make_username_from_full_name(full_name: str) -> str:
    base = _normalize_text(full_name).replace(" ", ".")
    base = base[:32] if base else "user"
    return base


def random_password(length: int = 16) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def bcrypt_laravel(password: str) -> str:
    import bcrypt  # lazy import

    salt = bcrypt.gensalt(rounds=10)
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")
    # Laravel/PHP commonly uses $2y$ prefix; PHP can also accept $2b$ on newer versions,
    # but to be safe we rewrite to $2y$.
    if hashed.startswith("$2b$"):
        hashed = "$2y$" + hashed[4:]
    return hashed


def build_mysql_url(prefix: str) -> str:
    host = _env(f"{prefix}_HOST")
    port = _env(f"{prefix}_PORT")
    db = _env(f"{prefix}_DATABASE")
    user = _env(f"{prefix}_USERNAME")
    pwd = os.getenv(f"{prefix}_PASSWORD", "")
    # pymysql works well on Windows
    return f"mysql+pymysql://{user}:{pwd}@{host}:{port}/{db}?charset=utf8mb4"


def get_table_columns(engine: Engine, table: str, schema: Optional[str] = None) -> List[str]:
    if schema is None:
        schema = engine.url.database
    q = text(
        """
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = :schema AND TABLE_NAME = :table
        ORDER BY ORDINAL_POSITION
        """
    )
    with engine.begin() as conn:
        rows = conn.execute(q, {"schema": schema, "table": table}).fetchall()
    return [r[0] for r in rows]


def select_one(engine: Engine, sql: str, params: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    with engine.begin() as conn:
        row = conn.execute(text(sql), params).mappings().first()
        return dict(row) if row else None


def ensure_cskh_status_code(engine4: Engine, code: str, expected_id: Optional[int] = None) -> Dict[str, Any]:
    row = select_one(
        engine4,
        "SELECT id, code, description FROM cskh_status WHERE code = :c LIMIT 1",
        {"c": code},
    )
    if not row:
        raise RuntimeError(f"Missing cskh_status.code='{code}'. Please create it before import.")
    if expected_id is not None and int(row.get("id") or 0) != expected_id:
        raise RuntimeError(
            f"cskh_status.code='{code}' exists but id={row.get('id')} (expected id={expected_id}). "
            "Fix the seed/data or remove the expected_id check."
        )
    return row


def upsert_customer_and_phone(
    engine4: Engine,
    customer_name: str,
    phone_raw: Any,
) -> Optional[int]:
    phone_norm = normalize_vn_phone(phone_raw)
    if not customer_name and not phone_norm:
        return None

    if phone_norm:
        existing = select_one(
            engine4,
            "SELECT customer_id FROM cskh_customer_phones WHERE phone_normalized = :p LIMIT 1",
            {"p": phone_norm},
        )
        if existing:
            return int(existing["customer_id"])

    customer_cols = set(get_table_columns(engine4, "cskh_customers"))

    payload: Dict[str, Any] = {}
    if "name" in customer_cols:
        payload["name"] = customer_name or phone_norm or "Khách hàng"
    if "created_at" in customer_cols:
        payload["created_at"] = datetime.now()
    if "updated_at" in customer_cols:
        payload["updated_at"] = datetime.now()

    if "name" not in payload:
        raise RuntimeError("cskh_customers table missing required column: name")

    columns = ", ".join(payload.keys())
    values = ", ".join([f":{k}" for k in payload.keys()])
    sql = f"INSERT INTO cskh_customers ({columns}) VALUES ({values})"

    with engine4.begin() as conn:
        res = conn.execute(text(sql), payload)
        customer_id = int(res.lastrowid)

        if phone_norm:
            # Insert phone; if another process inserted same phone concurrently, ignore.
            conn.execute(
                text(
                    """
                    INSERT IGNORE INTO cskh_customer_phones (customer_id, phone, phone_normalized, is_primary, created_at)
                    VALUES (:cid, :phone, :phone_norm, 1, NOW())
                    """
                ),
                {"cid": customer_id, "phone": str(phone_raw).strip(), "phone_norm": phone_norm},
            )

    return customer_id


def ensure_receiving_department(engine4: Engine, description: str) -> int:
    desc = (description or "").strip()
    if not desc:
        desc = "Kỹ thuật"
    found = select_one(
        engine4,
        "SELECT id FROM cskh_receiving_departments WHERE description = :d LIMIT 1",
        {"d": desc},
    )
    if found:
        return int(found["id"])
    dept_cols = set(get_table_columns(engine4, "cskh_receiving_departments"))
    payload: Dict[str, Any] = {"description": desc}
    if "is_active" in dept_cols:
        payload["is_active"] = 1
    if "created_at" in dept_cols:
        payload["created_at"] = datetime.now()
    if "updated_at" in dept_cols:
        payload["updated_at"] = datetime.now()

    columns = ", ".join(payload.keys())
    values = ", ".join([f":{k}" for k in payload.keys()])
    sql = f"INSERT INTO cskh_receiving_departments ({columns}) VALUES ({values})"

    with engine4.begin() as conn:
        res = conn.execute(text(sql), payload)
        return int(res.lastrowid)


def ensure_user_by_full_name(engine4: Engine, full_name: str) -> int:
    full_name = (full_name or "").strip()
    if not full_name:
        raise RuntimeError("Empty full_name for user")

    row = select_one(
        engine4,
        "SELECT id FROM users WHERE full_name = :n LIMIT 1",
        {"n": full_name},
    )
    if row:
        return int(row["id"])

    cols = set(get_table_columns(engine4, "users"))
    username_base = make_username_from_full_name(full_name)
    username = username_base

    # If username column exists, try to keep it unique.
    if "username" in cols:
        i = 1
        while True:
            exists = select_one(
                engine4,
                "SELECT id FROM users WHERE username = :u LIMIT 1",
                {"u": username},
            )
            if not exists:
                break
            i += 1
            username = f"{username_base}.{i}"[:64]

    pwd_plain = random_password()
    pwd_hash = bcrypt_laravel(pwd_plain)

    payload: Dict[str, Any] = {}
    if "full_name" in cols:
        payload["full_name"] = full_name
    if "name" in cols:
        payload["name"] = full_name
    if "username" in cols:
        payload["username"] = username
    if "email" in cols:
        # create a stable fake email if unique constraint exists
        local = re.sub(r"[^a-z0-9.]+", "", username.lower()) or "user"
        payload["email"] = f"{local}@import.local"
    if "password" in cols:
        payload["password"] = pwd_hash
    if "is_active" in cols:
        payload["is_active"] = 1
    if "created_at" in cols:
        payload["created_at"] = datetime.now()
    if "updated_at" in cols:
        payload["updated_at"] = datetime.now()

    # Minimal required fields fallback
    if not payload:
        raise RuntimeError("Users table columns not compatible with expected schema")

    columns = ", ".join(payload.keys())
    values = ", ".join([f":{k}" for k in payload.keys()])
    sql = f"INSERT INTO users ({columns}) VALUES ({values})"

    with engine4.begin() as conn:
        res = conn.execute(text(sql), payload)
        return int(res.lastrowid)


@dataclass
class ProductIndex:
    choices: List[str]
    meta: Dict[str, Dict[str, Any]]


def build_product_index(engine3: Engine) -> ProductIndex:
    sql = text(
        """
        SELECT id, product_name, name, model
        FROM products
        WHERE status IS NULL OR status != 0
        """
    )
    choices: List[str] = []
    meta: Dict[str, Dict[str, Any]] = {}
    with engine3.begin() as conn:
        rows = conn.execute(sql).mappings().all()
    for r in rows:
        pid = int(r["id"])
        fields = [r.get("product_name"), r.get("name"), r.get("model")]
        labels = [str(x).strip() for x in fields if x is not None and str(x).strip()]
        if not labels:
            continue
        # store a single canonical label for matching
        canonical = labels[0]
        key = f"{pid}:{canonical}"
        choices.append(key)
        meta[key] = {"id": pid, "canonical": canonical, "all_labels": labels}
    return ProductIndex(choices=choices, meta=meta)


def match_product(product_index: ProductIndex, raw_name: str) -> Tuple[Optional[int], Optional[str], int]:
    raw_name = (raw_name or "").strip()
    if not raw_name:
        return None, None, 0

    query = _normalize_text(raw_name)
    if not query:
        return None, None, 0

    # Custom scorer for rapidfuzz.process.extractOne.
    # Signature must be (query, choice, **kwargs), see rapidfuzz docs.
    def scorer(q: str, choice_key: str, **kwargs: Any) -> int:
        labels = product_index.meta[choice_key]["all_labels"]
        best = 0
        for lb in labels:
            lb_n = _normalize_text(lb)
            best = max(best, fuzz.token_set_ratio(q, lb_n))
        return best

    best = process.extractOne(query, product_index.choices, scorer=scorer)
    if not best:
        return None, None, 0
    choice_key, score, _idx = best
    if score < 70:
        return None, None, int(score)
    info = product_index.meta[choice_key]
    return int(info["id"]), str(info["canonical"]), int(score)


def parse_sheet_month_year(sheet_name: str) -> Tuple[int, int]:
    s = str(sheet_name or "").strip().replace(" ", "")
    m = re.match(r"^[Tt](\d{1,2})(\d{4})$", s)
    if not m:
        raise RuntimeError(
            f"Sheet name '{sheet_name}' không đúng định dạng. "
            "Ví dụ hợp lệ: T12025, T22025, T122025"
        )

    month = int(m.group(1))
    year = int(m.group(2))

    if month < 1 or month > 12:
        raise RuntimeError(f"Tháng trong sheet '{sheet_name}' không hợp lệ: {month}")

    return month, year


def parse_created_at(value: Any, sheet_month: int, sheet_year: int) -> Optional[datetime]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    if isinstance(value, datetime):
        return value

    # Nếu ô A chỉ là số ngày: 1, 2, 15, 31...
    if isinstance(value, (int, float)) and not pd.isna(value):
        day = int(value)
        if 1 <= day <= 31:
            return datetime(sheet_year, sheet_month, day, 0, 0, 0)

    s = str(value).strip()
    if not s:
        return None

    # Nếu là chuỗi chỉ chứa ngày, ví dụ "1", "15"
    if re.fullmatch(r"\d{1,2}", s):
        day = int(s)
        if 1 <= day <= 31:
            return datetime(sheet_year, sheet_month, day, 0, 0, 0)

    # Nếu là chuỗi kiểu "Ngày 15"
    m_day = re.search(r"(\d{1,2})", s)
    if m_day:
        day = int(m_day.group(1))
        if 1 <= day <= 31:
            return datetime(sheet_year, sheet_month, day, 0, 0, 0)

    # Nếu ô A đã là ngày đầy đủ dd/mm/yyyy thì vẫn parse bình thường
    try:
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            return None
        return ts.to_pydatetime()
    except Exception:
        return None


def _header_label(cell: Any) -> str:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return ""
    s = str(cell).strip()
    return "" if not s or s.lower() == "nan" else s


def classify_excel_header(label: str) -> Optional[str]:
    """
    Map dòng tiêu đề Excel (table.column) → khóa nội bộ.
    None = không dùng cho import có cấu trúc mới.
    """
    t = _header_label(label)
    if not t:
        return None
    low = re.sub(r"\s+", " ", t.lower())
    if low in ("cskh_tickets.created_at", "tickets.created_at"):
        return "created_at"
    if low == "cskh_status.code":
        return "status_code"
    if low == "cskh_customers.name":
        return "customer_name"
    if low == "cskh_customer_phones.phone":
        return "phone"
    if low == "cskh_tickets.issue_content":
        return "issue_content"
    if low == "cskh_tickets.handling":
        return "handling"
    if low == "cskh_solutions.resolution_text":
        return "resolution_text"
    if low == "users.full_name":
        return "receiver_name"
    if low == "cskh_tickets.source":
        return "source"
    if low == "cskh_receiving_departments.id":
        return "receiving_department_id"
    if low == "cskh_tickets.notes":
        return "notes"
    if low == "cskh_tickets.product_text":
        return "product_text_raw"
    n = _normalize_text(t)
    if "tinh" in n and "thanh" in n:
        return "province"
    return None


def build_column_map(df: pd.DataFrame) -> Tuple[Dict[str, int], bool]:
    """
    Đọc map cột từ tên cột (dòng 1 trong Excel khi header=0).
    Trả về (map khóa nội bộ -> chỉ số cột, header_mode).
    header_mode=True khi nhận diện đủ file export mới (có issue_content + created_at tối thiểu).
    """
    m: Dict[str, int] = {}
    for j, col in enumerate(df.columns):
        key = classify_excel_header(col)
        if key and key not in m:
            m[key] = j
    # Định dạng mới: có cột issue_content theo tên bảng (không còn E=product cố định).
    header_mode = "created_at" in m and "issue_content" in m and "customer_name" in m
    return m, header_mode


def parse_optional_int_id(value: Any) -> Optional[int]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float) and not pd.isna(value):
        if value == int(value):
            return int(value)
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return None
    if re.fullmatch(r"\d+", s):
        return int(s)
    try:
        f = float(s)
        if f == int(f):
            return int(f)
    except ValueError:
        pass
    return None


_LEGACY_COL_INDEX: Dict[str, int] = {
    "created_at": 0,
    "status_code": 1,
    "customer_name": 2,
    "phone": 3,
    "product_text_raw": 4,
    "issue_content": 5,
    "handling": 6,
    "receiver_name": 7,
    "source": 8,
}


def get_row_field(row: pd.Series, field: str, col_map: Dict[str, int], header_mode: bool) -> Any:
    if header_mode:
        j = col_map.get(field)
        if j is None:
            return None
        return row.iloc[j]
    j = _LEGACY_COL_INDEX.get(field)
    if j is None or j >= len(row):
        return None
    return row.iloc[j]


def str_cell(v: Any) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


# Map workflow status from sheet column B (TRẠNG THÁI) hoặc cskh_status.code
def map_workflow_status(sheet_status: str) -> str:
    raw = (sheet_status or "").strip()
    if not raw or raw.lower() == "nan":
        return "received"

    # Mã workflow trực tiếp (DB / API)
    code = re.sub(r"\s+", "", raw.lower())
    if code in ("received", "completed", "not_answered", "stopped"):
        return code

    # "!!!" = Dừng hỗ trợ (_normalize_text bỏ dấu ! nên phải check raw)
    compact = re.sub(r"\s+", "", raw)
    # Excel đôi khi dùng ký tự unicode như ❗/‼, hoặc user nhập "!! !"
    # Chỉ cần chuỗi sau khi bỏ khoảng trắng là toàn dấu chấm than (>=3 ký tự) thì coi là stopped.
    if (
        "!!!" in raw
        or compact == "!!!"
        or (len(compact) >= 3 and re.fullmatch(r"[!！‼❗]+", compact) is not None)
    ):
        return "stopped"

    s = _normalize_text(sheet_status)

    if "ko nghe may" in s or "khong nghe may" in s:
        return "not_answered"

    # Trạng thái hiển thị / mô tả (cskh_status) thường gặp
    if "hoan thanh" in s or s == "hoan thanh":
        return "completed"

    # Đã tiếp nhận (received): Chờ, Đang xử lý, Chưa LL, Lưu ý, Xử lý sau, ...
    if s == "cho" or s.startswith("cho "):
        return "received"
    if "dang xu ly" in s:
        return "received"
    if "chua ll" in s:
        return "received"
    if "luu y" in s:
        return "received"
    if "xu ly sau" in s:
        return "received"

    # OK = hoàn thành trên sheet
    if s == "ok":
        return "completed"

    # Nhãn lạ → received (tránh đổ hết vào completed, làm cột Kanban "Đã tiếp nhận" trống)
    return "received"


def insert_cskh_solution_resolution(
    conn: Any,
    solution_cols: set,
    ticket_id: int,
    resolution_text: str,
    created_at: Optional[datetime],
    workflow_status: str,
    completed_status_id: Optional[int],
) -> None:
    """Ghi `resolution_text` vào bảng cskh_solutions (nếu có nội dung)."""
    txt = (resolution_text or "").strip()
    if not txt:
        return
    if "ticket_id" not in solution_cols or "resolution_text" not in solution_cols:
        return

    payload: Dict[str, Any] = {
        "ticket_id": ticket_id,
        "resolution_text": txt,
    }
    if "created_at" in solution_cols:
        payload["created_at"] = created_at if created_at is not None else datetime.now()
    if "latest_reason" in solution_cols:
        payload["latest_reason"] = None
    if "status_id" in solution_cols and workflow_status == "completed" and completed_status_id is not None:
        payload["status_id"] = completed_status_id

    cols_i = list(payload.keys())
    col_sql = ", ".join(cols_i)
    val_sql = ", ".join([f":{c}" for c in cols_i])
    conn.execute(text(f"INSERT INTO cskh_solutions ({col_sql}) VALUES ({val_sql})"), payload)


def insert_ticket_activity_if_possible(
    engine4: Engine,
    ticket_id: int,
    user_id: int,
    action: str,
    type_: str,
    meta: Dict[str, Any],
    created_at: Optional[datetime],
) -> None:
    cols = set(get_table_columns(engine4, "cskh_ticket_activities"))
    payload: Dict[str, Any] = {}

    if "cskh_ticket_id" in cols:
        payload["cskh_ticket_id"] = ticket_id
    elif "ticket_id" in cols:
        payload["ticket_id"] = ticket_id
    else:
        return

    if "user_id" in cols:
        payload["user_id"] = user_id
    if "action" in cols:
        payload["action"] = action
    if "type" in cols:
        payload["type"] = type_
    if "meta" in cols:
        # Some MySQL drivers can't bind a Python dict directly.
        # Store as JSON string (compatible with JSON/TEXT columns).
        payload["meta"] = json.dumps(meta, ensure_ascii=False, default=str)
    if "created_at" in cols and created_at is not None:
        payload["created_at"] = created_at

    if not payload:
        return

    with engine4.begin() as conn:
        cols_i = list(payload.keys())
        col_sql = ", ".join(cols_i)
        val_sql = ", ".join([f":{c}" for c in cols_i])
        conn.execute(text(f"INSERT INTO cskh_ticket_activities ({col_sql}) VALUES ({val_sql})"), payload)


def import_excel(
    engine3: Engine,
    engine4: Engine,
    excel_path: str,
    sheet_name: Optional[str],
    start_row: int,
    dry_run: bool,
    fallback_created_by: Optional[int],
    tech_department_name: str,
) -> None:
    xls = pd.ExcelFile(excel_path)
    actual_sheet_name = sheet_name or xls.sheet_names[0]

    df = pd.read_excel(xls, sheet_name=actual_sheet_name, header=0)

    col_map, header_mode = build_column_map(df)
    print("Excel column mode:", "header (dòng 1 = table.column)" if header_mode else "legacy (A..I theo vị trí)")

    # Lấy tháng/năm từ tên sheet, ví dụ T12025 => tháng 1, năm 2025
    sheet_month, sheet_year = parse_sheet_month_year(actual_sheet_name)

    # Cột ngày bị merge trong Excel → ffill theo đúng cột `created_at`
    if header_mode:
        j_ct = col_map.get("created_at")
        if j_ct is not None:
            df.iloc[:, j_ct] = df.iloc[:, j_ct].ffill()
    else:
        if df.shape[1] < 9:
            raise RuntimeError(f"Excel định dạng cũ cần ít nhất 9 cột (A..I). Found: {df.shape[1]}")
        df.iloc[:, 0] = df.iloc[:, 0].ffill()

    product_index = build_product_index(engine3)

    imported = 0
    skipped = 0
    errors: List[str] = []

    # Validate workflow statuses exist in DB.
    ensure_cskh_status_code(engine4, "received", expected_id=1)
    ensure_cskh_status_code(engine4, "completed")
    ensure_cskh_status_code(engine4, "not_answered")
    ensure_cskh_status_code(engine4, "stopped")

    ticket_cols = set(get_table_columns(engine4, "cskh_tickets"))
    transfer_cols = set(get_table_columns(engine4, "cskh_ticket_transfers"))
    solution_cols = set(get_table_columns(engine4, "cskh_solutions"))
    tech_dept_id = ensure_receiving_department(engine4, tech_department_name)

    completed_row = select_one(
        engine4,
        "SELECT id FROM cskh_status WHERE code = 'completed' LIMIT 1",
        {},
    )
    completed_status_id = int(completed_row["id"]) if completed_row else None

    for i in range(start_row, len(df)):
        row = df.iloc[i]

        created_at = parse_created_at(
            get_row_field(row, "created_at", col_map, header_mode),
            sheet_month,
            sheet_year,
        )
        sheet_status = str_cell(get_row_field(row, "status_code", col_map, header_mode))
        customer_name = str_cell(get_row_field(row, "customer_name", col_map, header_mode))
        phone_raw = get_row_field(row, "phone", col_map, header_mode)
        customer_name, phone_raw, extra_contact_note = apply_customer_column_split_for_import(
            customer_name, phone_raw
        )
        product_raw = str_cell(get_row_field(row, "product_text_raw", col_map, header_mode))
        issue_content = str_cell(get_row_field(row, "issue_content", col_map, header_mode))
        handling_for_ticket = str_cell(get_row_field(row, "handling", col_map, header_mode)) if header_mode else ""
        resolution_for_solution = (
            str_cell(get_row_field(row, "resolution_text", col_map, header_mode))
            if header_mode
            else str_cell(get_row_field(row, "handling", col_map, header_mode))
        )
        receiver_name = str_cell(get_row_field(row, "receiver_name", col_map, header_mode))
        source = str_cell(get_row_field(row, "source", col_map, header_mode))
        notes_excel = str_cell(get_row_field(row, "notes", col_map, header_mode))
        province_txt = str_cell(get_row_field(row, "province", col_map, header_mode))
        sheet_dept_id = (
            parse_optional_int_id(get_row_field(row, "receiving_department_id", col_map, header_mode))
            if header_mode
            else None
        )

        nv = normalize_vn_phone(phone_raw)
        if not any(
            [
                created_at,
                customer_name,
                nv,
                product_raw,
                issue_content,
                handling_for_ticket,
                resolution_for_solution,
                receiver_name,
                source,
                notes_excel,
                province_txt,
            ]
        ):
            skipped += 1
            continue

        try:
            customer_id = upsert_customer_and_phone(engine4, customer_name, phone_raw)

            # Status mapping (cột trạng thái / cskh_status.code): xem map_workflow_status()
            workflow_status = map_workflow_status(sheet_status)

            product_id, product_canonical, product_score = match_product(product_index, product_raw)

            notes_lines: List[str] = []
            if notes_excel:
                notes_lines.append(notes_excel)
            if extra_contact_note:
                notes_lines.append(extra_contact_note)
            if province_txt:
                notes_lines.append(f"Tỉnh/Thành: {province_txt}")
            if not header_mode and sheet_status:
                notes_lines.append(f"Trạng thái sheet: {sheet_status}")
            notes = "\n".join(notes_lines) if notes_lines else None

            is_tech = _normalize_text(receiver_name) == _normalize_text("Kỹ thuật")

            created_by: Optional[int] = None
            if is_tech:
                created_by = fallback_created_by
            else:
                if receiver_name:
                    created_by = ensure_user_by_full_name(engine4, receiver_name)
                else:
                    created_by = fallback_created_by

            if created_by is None:
                raise RuntimeError("Cannot resolve created_by (receiver missing and no fallback)")

            ticket_payload: Dict[str, Any] = {}
            if "customer_id" in ticket_cols:
                ticket_payload["customer_id"] = customer_id
            if "workflow_status" in ticket_cols:
                ticket_payload["workflow_status"] = workflow_status
            if "product_text" in ticket_cols:
                # Sheet không có sản phẩm → NULL; có sản phẩm + match → tên chuẩn; có nhưng không match → NULL
                if not product_raw:
                    ticket_payload["product_text"] = None
                else:
                    ticket_payload["product_text"] = product_canonical if product_id else None
            if "product_id" in ticket_cols:
                ticket_payload["product_id"] = int(product_id) if product_id is not None else None
            if "issue_content" in ticket_cols:
                ticket_payload["issue_content"] = issue_content or None
            if "handling" in ticket_cols:
                ticket_payload["handling"] = (handling_for_ticket or None) if header_mode else None
            if "source" in ticket_cols:
                ticket_payload["source"] = (source[:32] if source else None)
            if "notes" in ticket_cols:
                ticket_payload["notes"] = notes
            if "created_by" in ticket_cols:
                ticket_payload["created_by"] = created_by
            if "receiving_department_id" in ticket_cols:
                if sheet_dept_id is not None:
                    ticket_payload["receiving_department_id"] = sheet_dept_id
                elif is_tech:
                    ticket_payload["receiving_department_id"] = tech_dept_id
                else:
                    ticket_payload["receiving_department_id"] = None
            if "current_department_id" in ticket_cols:
                if sheet_dept_id is not None:
                    ticket_payload["current_department_id"] = sheet_dept_id
                elif is_tech:
                    ticket_payload["current_department_id"] = tech_dept_id
                else:
                    ticket_payload["current_department_id"] = None
            if "created_at" in ticket_cols:
                ticket_payload["created_at"] = created_at
            if "updated_at" in ticket_cols:
                ticket_payload["updated_at"] = created_at or datetime.now()

            # Ensure minimal required columns exist
            if "created_by" in ticket_cols and "created_by" not in ticket_payload:
                raise RuntimeError("cskh_tickets requires created_by but script couldn't set it")
            if "workflow_status" in ticket_cols and "workflow_status" not in ticket_payload:
                raise RuntimeError("cskh_tickets requires workflow_status but script couldn't set it")

            if dry_run:
                imported += 1
                continue

            with engine4.begin() as conn:
                cols = list(ticket_payload.keys())
                col_sql = ", ".join(cols)
                val_sql = ", ".join([f":{c}" for c in cols])
                res = conn.execute(text(f"INSERT INTO cskh_tickets ({col_sql}) VALUES ({val_sql})"), ticket_payload)
                ticket_id = int(res.lastrowid)

                insert_cskh_solution_resolution(
                    conn,
                    solution_cols,
                    ticket_id,
                    resolution_for_solution,
                    created_at,
                    workflow_status,
                    completed_status_id,
                )

                activity_meta: Dict[str, Any] = {
                    "import": "excel",
                    "header_mode": header_mode,
                    "sheet_status": sheet_status or None,
                    "product_raw": product_raw or None,
                    "product_id": product_id,
                    "product_name": product_canonical,
                    "product_score": product_score,
                    "source": source or None,
                    "resolution_text": resolution_for_solution or None,
                    "handling": (handling_for_ticket or None) if header_mode else None,
                }
                insert_ticket_activity_if_possible(
                    engine4=engine4,
                    ticket_id=ticket_id,
                    user_id=created_by,
                    action="import",
                    type_="excel",
                    meta=activity_meta,
                    created_at=created_at,
                )

                if is_tech:
                    transfer_payload: Dict[str, Any] = {}
                    if "ticket_id" in transfer_cols:
                        transfer_payload["ticket_id"] = ticket_id
                    if "to_department_id" in transfer_cols:
                        transfer_payload["to_department_id"] = tech_dept_id
                    if "error_encountered" in transfer_cols:
                        transfer_payload["error_encountered"] = "Chuyển tiếp từ import Excel"
                    if "descriptions" in transfer_cols:
                        transfer_payload["descriptions"] = (issue_content or "")[:2000]
                    if "note" in transfer_cols:
                        transfer_payload["note"] = "Auto transfer: receiver=Kỹ thuật"
                    if "transferred_by" in transfer_cols:
                        transfer_payload["transferred_by"] = created_by
                    if "created_at" in transfer_cols:
                        transfer_payload["created_at"] = created_at

                    cols_t = list(transfer_payload.keys())
                    col_sql_t = ", ".join(cols_t)
                    val_sql_t = ", ".join([f":{c}" for c in cols_t])
                    conn.execute(text(f"INSERT INTO cskh_ticket_transfers ({col_sql_t}) VALUES ({val_sql_t})"), transfer_payload)

            imported += 1

        except Exception as e:
            errors.append(f"Row {i+2}: {e}")  # +2 because header=0 and 1-indexed Excel rows

    print(f"Imported rows: {imported}")
    print(f"Skipped empty rows: {skipped}")
    if errors:
        print(f"Errors: {len(errors)}")
        for msg in errors[:50]:
            print(msg)
        if len(errors) > 50:
            print(f"... {len(errors)-50} more")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import CSKH tickets from Excel to mysql4.cskh_tickets (map products from mysql3.products).")
    parser.add_argument("--excel", required=True, help="Path to Excel file")
    parser.add_argument(
        "--normalize-excel-out",
        default=None,
        metavar="OUT.xlsx",
        help="Chuẩn hóa file: tách tên/SĐT, ghi SĐT vào cột phone nếu trống; sau đó (mặc định) gộp ô cột ngày tạo trên cùng file out. Không kết nối DB.",
    )
    parser.add_argument(
        "--no-merge-created-at",
        action="store_true",
        help="Dùng với --normalize-excel-out: chỉ tách tên/SĐT, không gộp cột created_at.",
    )
    parser.add_argument(
        "--merge-created-at-out",
        default=None,
        metavar="OUT.xlsx",
        help="Gộp ô dọc cột ngày tạo (created_at / cột A legacy): các dòng trống dưới cùng một ngày (ffill) merge như Excel. Không kết nối DB.",
    )
    parser.add_argument("--sheet", default=None, help="Excel sheet name (default: first sheet)")
    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help="Import từng sheet trong file; sheet tên không đúng dạng T{tháng}{năm} (vd T12025) sẽ bị bỏ qua",
    )
    parser.add_argument("--start-row", type=int, default=0, help="0-based data row index (excluding header). Default 0.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate only; do not write to DB")
    parser.add_argument("--fallback-created-by", type=int, default=None, help="User id to use when receiver is empty or receiver is 'Kỹ thuật'")
    parser.add_argument("--tech-department-name", default="Kỹ thuật", help="cskh_receiving_departments.description to use/create for 'Kỹ thuật'")
    args = parser.parse_args()

    if args.all_sheets and args.sheet is not None:
        raise SystemExit("Chỉ dùng một trong hai: --all-sheets hoặc --sheet, không dùng cùng lúc.")

    if int(bool(args.normalize_excel_out)) + int(bool(args.merge_created_at_out)) > 1:
        raise SystemExit("Chỉ dùng một trong: --normalize-excel-out hoặc --merge-created-at-out.")

    if args.normalize_excel_out:
        normalize_excel_customer_columns(
            excel_path=args.excel,
            sheet_name=args.sheet,
            out_path=args.normalize_excel_out,
            all_sheets=args.all_sheets,
            also_merge_created_at=not args.no_merge_created_at,
        )
        if args.no_merge_created_at:
            print(f"Đã ghi file chuẩn hóa (chưa gộp cột ngày): {args.normalize_excel_out}")
        else:
            print(f"Đã ghi file chuẩn hóa + đã gộp cột ngày tạo: {args.normalize_excel_out}")
        return

    if args.merge_created_at_out:
        merge_excel_created_at_column(
            excel_path=args.excel,
            sheet_name=args.sheet,
            out_path=args.merge_created_at_out,
            all_sheets=args.all_sheets,
        )
        print(f"Đã ghi file đã gộp cột ngày: {args.merge_created_at_out}")
        return

    load_dotenv()

    engine3 = create_engine(build_mysql_url("DB3"), pool_pre_ping=True, future=True)
    engine4 = create_engine(build_mysql_url("DB4"), pool_pre_ping=True, future=True)

    def run_one(sheet: Optional[str]) -> None:
        import_excel(
            engine3=engine3,
            engine4=engine4,
            excel_path=args.excel,
            sheet_name=sheet,
            start_row=args.start_row,
            dry_run=args.dry_run,
            fallback_created_by=args.fallback_created_by,
            tech_department_name=args.tech_department_name,
        )

    if args.all_sheets:
        xls = pd.ExcelFile(args.excel)
        for sn in xls.sheet_names:
            print(f"\n================= Sheet: {sn!r} ==================")
            try:
                run_one(sn)
            except RuntimeError as e:
                msg = str(e)
                if "không đúng định dạng" in msg or "không hợp lệ" in msg:
                    print(f"SKIP sheet: {msg}")
                    continue
                raise
    else:
        run_one(args.sheet)


if __name__ == "__main__":
    main()

